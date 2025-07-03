import sys
import os
import pandas as pd
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QFileDialog,
    QListWidget,
    QVBoxLayout,
    QPushButton,
    QWidget,
    QMessageBox,
    QLabel,
    QLineEdit,
    QProgressBar,
    QAbstractItemView
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


from PyQt5.QtCore import Qt, QThread, pyqtSignal,QObject
import re
import time as today
from datetime import datetime
import time
import zipfile
from collections import Counter
from queue import Queue
from PyQt5.QtWidgets import QLineEdit  # Add this import statement


from PyQt5.QtCore import QThread, pyqtSignal
import os
import re
import time
import mmap
from concurrent.futures import ProcessPoolExecutor, as_completed

# Move this function to top-level so it can be pickled by multiprocessing
def process_single_log(args):
    filename, folder_path, selected_file = args
    start_time = time.time()
    filepath = os.path.join(folder_path, filename)
    node_log = re.search(r"^(.*?).log$", filename, re.IGNORECASE).group(1) if re.search(r"^(.*?).log$", filename, re.IGNORECASE) else filename
    local_log_data = []

    # Compile regex patterns once
    pattern_cmd = re.compile(r"^[A-Z0-9\_\-]{4,180}\>(.*?)$", re.IGNORECASE)
    pattern_run_script = re.compile(r".*?run\s+.*?\$nodename\_(.*?).mos$", re.IGNORECASE)
    pattern_truni = re.compile(r"^(trun|truni)\s+(.*?)$", re.IGNORECASE)
    pattern_tag = re.compile(r'^!!!!.*?TAG\s+:"(.*?)".*?$', re.IGNORECASE)
    pattern_alt_tag = re.compile(r"^>>>\s+(\[.*?\]).*?$", re.IGNORECASE)
    pattern_set = re.compile(r"^(SET\s+.*?)$", re.IGNORECASE)
    pattern_create = re.compile(r"^(CREATE.*?)$", re.IGNORECASE)
    pattern_delete = re.compile(r"^(DELETE.*?)$", re.IGNORECASE)
    tag_errs = [
        re.compile(r"^(ERROR:.*?:.*?)$", re.IGNORECASE),
        re.compile(r"(!!!!\s+(ERROR|Processing):.*?)$", re.IGNORECASE),
        re.compile(r"(!!!!\s+Processing.*?)$", re.IGNORECASE),
        re.compile(r"^(>>>.*?:.*?)$", re.IGNORECASE),
        re.compile(r"^(Total.*?MOs\s+attempted.*?)$", re.IGNORECASE)
    ]

    try:
        with open(filepath, 'r', encoding='utf-8', errors='ignore') as file:
            mm = mmap.mmap(file.fileno(), 0, access=mmap.ACCESS_READ)
            type_script = "NULL"
            type_script_rnc = "NULL"
            line_execute = "NULL"
            TAG_RESULT = "NULL"
            truni_script = "NULL"
            TAG_RESULT_FULL = "NULL"
            print_tag = "NULL"
            number_tag = 0
            cmd_get = "NULL"
            att_list = []

            for line in iter(mm.readline, b''):
                line = line.decode(errors='ignore')
                cmd_match = pattern_cmd.search(line)
                cmd_get_1 = cmd_match.group(1).strip() if cmd_match else line_execute
                # --- RNC CR SCRIPT (truni/trun) ---
                type_script = re.search(r".*?run\s+.*?\$nodename\_(.*?).mos$", line, re.IGNORECASE).group(1) if re.search(r".*?run\s+.*?\$nodename\_(.*?).mos$", line, re.IGNORECASE) else type_script
                truni_script = re.search(r".*?(trun|truni)\s+(.*?)$", line, re.IGNORECASE).group(1) if re.search(r".*?(trun|truni)\s+(.*?)$", line, re.IGNORECASE) else truni_script
                if truni_script != "NULL":
                    type_script_rnc = (m.group(2) if (m := re.search(r".*?trun(i)?\s+(.*?).mo(s)?$", line, re.IGNORECASE)) else type_script_rnc)
                    line_execute = re.search(r"^(CREATE.*?)$", line, re.IGNORECASE).group(1) if re.search(r"^(CREATE.*?)$", line, re.IGNORECASE) else line_execute
                    line_execute = re.search(r"^(DELETE.*?)$", line, re.IGNORECASE).group(1) if re.search(r"^(DELETE.*?)$", line, re.IGNORECASE) else line_execute
                    line_execute = re.search(r"^(SET\s+.*?)$", line, re.IGNORECASE).group(1) if re.search(r"^(SET\s+.*?)$", line, re.IGNORECASE) else line_execute
                    if re.search(r"^!!!!.*?TAG\s+:\"(.*?)\".*?$", line, re.IGNORECASE):
                        TAG_RESULT = re.search(r"^!!!!.*?TAG\s+:\"(.*?)\".*?$", line, re.IGNORECASE).group(1)
                        TAG_RESULT_FULL = line.rstrip()
                    else:

                        
                        TAG_RESULT = TAG_RESULT
                        TAG_RESULT_FULL = TAG_RESULT_FULL
                    if re.search(r"^>>>\s+(\[.*?\]).*?$", line, re.IGNORECASE):
                        TAG_RESULT = re.search(r"^>>>\s+(\[.*?\]).*?$", line, re.IGNORECASE).group(1)
                        TAG_RESULT_FULL = line.rstrip()
                    else:
                        TAG_RESULT = TAG_RESULT
                        TAG_RESULT_FULL = TAG_RESULT_FULL
                    if (len(line.rstrip()) == 0 and line_execute != "NULL"):
                        if(len(TAG_RESULT_FULL.rstrip()) == 0):
                            TAG_RESULT = "Executed"
                        TAG_REPORT, TAG_COLOR = CATEGORY_CHECKING1(TAG_RESULT)
                        att_list = []
                        att_list.append(TAG_RESULT_FULL.strip())
                        local_log_data.append((selected_file, node_log, type_script_rnc, line_execute.strip(), TAG_REPORT, TAG_RESULT, att_list))
                        line_execute = "NULL"
                        TAG_RESULT = "NULL"
                        TAG_RESULT_FULL = ""
                # --- END RNC CR SCRIPT (truni/trun) ---
                # The rest of the original logic (for BB CR SCRIPT and others) remains unchanged from before.
                if truni_script == "NULL":
                    if pattern_cmd.match(line):
                        if number_tag == 1:
                            if re.search(r"^(del|rdel|crn|set)", cmd_get.strip(), re.IGNORECASE):
                                TAG_REPORT, TAG_COLOR = CATEGORY_CHECKING1(TAG_RESULT)
                                local_log_data.append((selected_file, node_log, type_script, line_execute.strip(), TAG_REPORT.strip(), TAG_RESULT.strip(), att_list))                                
                                
                                ###local_log_data.append((selected_file ,node_log,type_script, line_execute.strip(),TAG_REPORT.strip() ,TAG_RESULT.strip() , att_list    ))  # Store filename and line
                                TAG_RESULT = "NULL"
                                print_tag = "NULL"
                                number_tag = 0
                            else:
                                number_tag = 0
                        number_tag += 1
                        cmd_get = cmd_get_1

                    for pattern in tag_errs:
                        match = pattern.search(line)
                        if match:
                            TAG_RESULT = match.group(1)

                    if pattern_cmd.match(line):
                        att_list = ["", line.strip()]
                    else:
                        att_list.append(line.strip())

                if "Checking ip contact...Not OK" in line:
                    local_log_data.append((selected_file, node_log, "", "", "UNREMOTE", line.strip(), ""))
                if re.search(r"(?i)tbac\s*control\s*-\s*unauthori[sz]ed\s*network\s*element", line):
                    local_log_data.append((selected_file, node_log, "", "", "UNREMOTE", line.strip(), ""))



    except Exception as e:
        print(f"Error processing {filename}: {e}")
    end_time = time.time()
    print(f"Processed {filename} in {end_time - start_time:.2f}s")

    # --- New Section: Parse LOG_Alarm_bf, LOG_status_bf, LOG_Alarm_af, LOG_status_af ---
    from io import StringIO

    log_data_sections = {
        'LOG_Alarm_bf': [],
        'LOG_status_bf': [],
        'LOG_Alarm_af': [],
        'LOG_status_af': [],
    }

    mode = None
    collecting = False

    with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
        lines = f.readlines()

    for line in lines:
        stripped = line.strip()
        if '####LOG_Alarm_bf' in stripped:
            mode, collecting = 'LOG_Alarm_bf', False
            continue
        elif '####LOG_status_bf' in stripped:
            mode, collecting = 'LOG_status_bf', False
            continue
        elif '####LOG_Alarm_af' in stripped:
            mode, collecting = 'LOG_Alarm_af', False
            continue
        elif '####LOG_status_af' in stripped:
            mode, collecting = 'LOG_status_af', False
            continue

        if stripped == "" and collecting:
            mode, collecting = None, False
            continue

        if ';' in stripped and mode:
            collecting = True
            log_data_sections[mode].append(stripped)

    nodename = os.path.splitext(os.path.basename(filepath))[0]

    def parse_alarm(data_lines):
        if not data_lines:
            return pd.DataFrame()
        df = pd.read_csv(StringIO('\n'.join(data_lines)), sep=';', engine='python')
        df.columns = [col.strip() for col in df.columns]  # Strip spaces from headers
        df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        df['NODENAME'] = nodename
        return df

    def parse_status(data_lines):
        if not data_lines:
            return pd.DataFrame()
        headers = [l.split(';') for l in data_lines if l.startswith("MO")]
        if not headers:
            return pd.DataFrame()
        header = [h.strip() for h in headers[-1]]  # Strip header names
        rows = [
            [cell.strip() for cell in l.split(';')]
            for l in data_lines
            if not l.startswith("MO") and len(l.split(';')) == len(header)
        ]
        df = pd.DataFrame(rows, columns=header)
        df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
        df['NODENAME'] = nodename
        return df



    df_LOG_Alarm_bf = parse_alarm(log_data_sections['LOG_Alarm_bf'])
    df_LOG_Alarm_af = parse_alarm(log_data_sections['LOG_Alarm_af'])
    df_LOG_status_bf = parse_status(log_data_sections['LOG_status_bf'])
    df_LOG_status_af = parse_status(log_data_sections['LOG_status_af'])

    return {
        "log_data": local_log_data,
        "df_LOG_Alarm_bf": df_LOG_Alarm_bf,
        "df_LOG_Alarm_af": df_LOG_Alarm_af,
        "df_LOG_status_bf": df_LOG_status_bf,
        "df_LOG_status_af": df_LOG_status_af,
    }




def CATEGORY_CHECKING1(TAG_RESULT):
    # Check if the string contains certain substrings
    if "Proxy ID" in TAG_RESULT:
        TAG_REMARK = "1 MOs created"
        TAG_COLOR="42FF00"
    elif TAG_RESULT == "Executed" or re.search(r"([0-9]{1,90}\s+.*?\-\-\>.*?set\.)$", TAG_RESULT, re.IGNORECASE):
        TAG_REMARK = "1 MOs set"
        TAG_COLOR="42FF00" 
    elif TAG_RESULT == "Mo deleted":
        TAG_REMARK = "Mo deleted"
        TAG_COLOR="42FF00"             
    elif "MO already exists" in TAG_RESULT or TAG_RESULT == "MoNameAlreadyTaken":
        TAG_REMARK = "1 MOs already exists"
        TAG_COLOR="FF7575"
    elif "Parent MO not found" in TAG_RESULT:
        TAG_REMARK = "Parent MO not found"
        TAG_COLOR="FF7575"
    elif TAG_RESULT == "MoNotFound":
        TAG_REMARK = "MO not found"
        TAG_COLOR="FF7575"  
    elif re.search(r"operation\-failed", TAG_RESULT, re.IGNORECASE) or re.search(r"unknown\-attribute", TAG_RESULT, re.IGNORECASE) or "failure" in TAG_RESULT:
        TAG_REMARK = "ERROR operation-failed"
        TAG_COLOR="FF7575"             
    else:
        TAG_REMARK = TAG_RESULT 
        TAG_COLOR="FF7575"
            
    return  TAG_REMARK, TAG_COLOR 
    
    
    
def CATEGORY_CHECKING(TAG_RESULT):
    # Check if the string contains certain substrings
    check_number_executed = int(re.search(r"^Total.*?MOs\s+attempted\,\s+([0-9]{1,5})\s+MOs.*?$", TAG_RESULT, re.IGNORECASE).group(1)) if re.search(r"^(Total.*?MOs\s+attempted.*?)$", TAG_RESULT, re.IGNORECASE) else "NULL"
    Action_GET = re.search(r"^Total.*?MOs\s+attempted\,\s+[0-9]{1,5}\s+MOs\s+(.*?)$", TAG_RESULT, re.IGNORECASE).group(1) if re.search(r"^(Total.*?MOs\s+attempted.*?)$", TAG_RESULT, re.IGNORECASE) else "NULL"
    
    if TAG_RESULT == "1 MOs created" or TAG_RESULT == "1 MOs set" or TAG_RESULT == "Mo deleted":
        TAG_COLOR="42FF00"
    elif re.search(r"^Total.*?MOs\s+attempted", TAG_RESULT, re.IGNORECASE) and check_number_executed > 0:
        TAG_COLOR="42FF00"
    else:
        TAG_COLOR="FF7575"
    
    ###TOTAL X 
    if "TOTAL X" in TAG_RESULT:
        TAG_COLOR="42FF00"

    if (re.search(r"^Total.*?MOs\s+attempted", TAG_RESULT, re.IGNORECASE) and check_number_executed > 0):
        MSG = "Success"
    else:
        MSG = "Failed"
    

        
        
    if re.search(r"^Total.*?MOs\s+attempted", TAG_RESULT, re.IGNORECASE):
        TAG_RETURN = f'''{MSG} to {Action_GET}'''
    else:
        TAG_RETURN = TAG_RESULT
            
    return  TAG_RETURN, TAG_COLOR      


        
class ExcelReaderApp(QMainWindow):
    processing_finished = pyqtSignal()
    
    def __init__(self, start_path=None):
        super().__init__()
        self.start_path = start_path or os.path.expanduser('~')
        self.initUI()
        self.file_queue = Queue()  # Queue to store selected files
        self.append_log_data = []  # Initialize an empty list to store log data
        self.date_report = today.strftime('%Y%m%d_%H%M%S')

    
    def initUI(self):
        self.setWindowTitle("REPORT CR GENERATOR")
        self.setGeometry(100, 100, 400, 300)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        
        layout = QVBoxLayout(self.central_widget)

        placeholder_style = "font-style: italic;"
        self.browse_button = QPushButton("Browse", self)
        layout.addWidget(self.browse_button)
        self.browse_button.clicked.connect(self.open_folder_dialog)

        # Allow multi-selection for folders
        self.file_list = QListWidget(self)
        self.file_list.setSelectionMode(QAbstractItemView.MultiSelection)
        # Style selected items: white, bold text and distinct background
        self.file_list.setStyleSheet('''
            QListWidget::item:selected {
                background: #0078d7;
                color: white;
                font-weight: bold;
            }
            QListWidget::item {
                color: black;
            }
        ''')
        layout.addWidget(self.file_list)
        
        # Add text box input
        self.input_directory = QLineEdit(self)
        self.input_directory.setPlaceholderText("Input Directory to upload script on ENM: e.g: /home/shared/username/sample_folder or ~/sample_folder")
        layout.addWidget(self.input_directory)        

        self.read_button = QPushButton("Generate Report", self)
        layout.addWidget(self.read_button)
        self.read_button.clicked.connect(self.read_selected_excel)

        self.quit_button = QPushButton("Quit", self)
        layout.addWidget(self.quit_button)
        self.quit_button.clicked.connect(self.close)

        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        # Add phase label
        self.phase_label = QLabel("Ready")
        self.phase_label.setStyleSheet("color: white; font-weight: bold;")
        layout.addWidget(self.phase_label)
        # Add details label
        self.details_label = QLabel("")
        self.details_label.setStyleSheet("color: white; font-weight: bold;")
        layout.addWidget(self.details_label)

        self.show()


    def show_success_message(self, message):
        success_box = QMessageBox()
        success_box.setIcon(QMessageBox.Information)
        success_box.setWindowTitle("Success")
        success_box.setText(message)
        success_box.exec_()

    def show_error_message(self, message):
        error_box = QMessageBox()
        error_box.setIcon(QMessageBox.Critical)
        error_box.setWindowTitle("Error")
        error_box.setText(message)
        error_box.exec_()


    def update_overall_progress(self, value):
        print(f"[DEBUG] Progress bar update: {value}")
        self.progress_bar.setValue(value)

    def on_thread_finished(self, file_path, log_data , selected_file , output_dir):
        report_file = f"{output_dir}\{selected_file}.xlsx"
        print("[DEBUG] About to start ExcelWriterThread")
        self.writer_thread = ExcelWriterThread(log_data, report_file, selected_file)
        self.writer_thread.progress_changed.connect(self.update_overall_progress)
        self.writer_thread.phase_changed.connect(self.update_phase_label)
        self.writer_thread.details_changed.connect(self.update_details_label)
        def on_file_written(_):
            print(f"Your Report Available On :\n{report_file}\n\n")
            self.append_log_data.extend(log_data)
            # Check if all files are processed
            if self.file_queue.empty():
                self.processing_finished.emit()  # Signal that all files are processed
                work_dir = os.getcwd()
                self.summary_excel = os.path.join(work_dir, "output", "Summary_" + self.date_report + ".xlsx")
                # Use ExcelWriterThread for summary report (all heavy work in thread)
                self.summary_writer_thread = ExcelWriterThread(self.append_log_data, self.summary_excel, "NONAME")
                self.summary_writer_thread.progress_changed.connect(self.update_overall_progress)
                self.summary_writer_thread.phase_changed.connect(self.update_phase_label)
                self.summary_writer_thread.details_changed.connect(self.update_details_label)
                def on_summary_written(_):
                    # Now you can proceed with further logic like saving or parsing summary
                    self.parse_summary(pd.DataFrame(self.append_log_data, columns=["CR", "Site", "Script", "Command", "Report", "TAG", "Full_cmd"]), self.summary_excel)
                    self.show_success_message("ALL DONE")
                self.summary_writer_thread.finished.connect(on_summary_written)
                self.summary_writer_thread.start()
        self.writer_thread.finished.connect(on_file_written)
        self.writer_thread.start()

    def parse_summary(self, pivot_df, excel_temp):
        workbook = openpyxl.load_workbook(excel_temp)
        pivot_sheet = workbook.create_sheet(title="Pivot Table")
        rows = list(dataframe_to_rows(pivot_df, index=True, header=True))

        # Set column widths
        column_widths = {2: 60, 3: 30, 4: 12, 5: 6, 6: 12}
        for col_idx, width in column_widths.items():
            pivot_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Write rows to worksheet with progress bar
        for r_idx, row in enumerate(tqdm(rows, desc="Creating Pivot Table", unit="row"), 1):
            for c_idx, value in enumerate(row, 1):
                cell = pivot_sheet.cell(row=r_idx, column=c_idx, value=value)

                # Apply bold font to index column
                if c_idx == 2:
                    cell.font = Font(bold=True)

                # Apply color fill to specific category
                if c_idx == 3:
                    TAG_REMARK, TAG_COLOR = CATEGORY_CHECKING(value)
                    cell.fill = PatternFill(start_color=TAG_COLOR, end_color=TAG_COLOR, fill_type="solid")

                # Apply border to all cells
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

        # Highlight the header row (first row)
        for col_idx in range(1, 7):
            pivot_sheet.cell(row=1, column=col_idx).fill = PatternFill(start_color="0EA1DD", end_color="0EA1DD", fill_type="solid")

        # Optionally delete second row (as in original)
        pivot_sheet.delete_rows(2)

        # Save the workbook
        workbook.save(excel_temp)
   

    def open_folder_dialog(self):
        self.folder_path = QFileDialog.getExistingDirectory(self, "Select Folder", self.start_path)
        if self.folder_path:
            self.populate_file_list()
    
    def populate_file_list(self):
        self.file_list.clear()
        if self.folder_path:
            folders = [
                folder
                for folder in os.listdir(self.folder_path)
                if os.path.isdir(os.path.join(self.folder_path, folder))
            ]
            self.file_list.addItems(folders)
        


     


    def read_selected_excel(self):
        self.file_queue.queue.clear()  # Clear existing queue

        selected_items = self.file_list.selectedItems()
        if not selected_items:
            return

        # Queue all selected folders for processing
        for item in selected_items:
            self.file_queue.put(item.text())

        # Set progress bar to 1% to show progress is ongoing before any log reading starts
        self.progress_bar.setValue(1)
        # Start processing files (log reading is already threaded)
        self.process_next_file()

    def process_next_file(self):
        if not self.file_queue.empty():
            selected_file = self.file_queue.get()
            file_path = os.path.join(self.folder_path)
            work_dir = os.getcwd()
            folder_enm = self.input_directory.text()
            self.output_dir = os.path.join(work_dir, "output")
            self.result_dir = os.path.join(work_dir, "result")
            try:
                self.progress_bar.setValue(0)
                self.worker_thread = WorkerThread(file_path, selected_file, self.output_dir)
                self.worker_thread.overall_progress.connect(self.update_overall_progress)
                self.worker_thread.phase_changed.connect(self.update_phase_label)
                self.worker_thread.details_changed.connect(self.update_details_label)
                self.worker_thread.finished.connect(self.on_thread_finished)
                self.worker_thread.finished.connect(self.process_next_file)
                self.worker_thread.start()
            except Exception as e:
                print(f"Error reading Excel file: {e}")
                self.show_error_message(f"Error reading Excel file: {e}")
        #####else:
        #####    self.processing_finished.emit()  # Signal that all files are processed
        #####    work_dir = os.getcwd()
        #####    self.summary_excel = os.path.join(work_dir, "output", "SUM.xlsx")
        #####    write_logs_to_excel(self.append_log_data, self.summary_excel, "NONAME")
        #####    self.show_success_message(f"ALL DONE")
        #####    ##print(self.append_log_data)
            
            

    # Existing methods...

    
    def check_folder (self,output_dir):
        isExist_output_dir = os.path.exists(output_dir)
                            
        if not isExist_output_dir:
            os.makedirs(output_dir)
        return output_dir

    




from tqdm import tqdm

# Function to write log data to an Excel file
def write_logs_to_excel(log_data, excel_filename, selected_file, progress_callback=None, df_LOG_Alarm_bf=None, df_LOG_Alarm_af=None, df_LOG_status_bf=None, df_LOG_status_af=None):
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active

    # Define a regular expression pattern to remove illegal characters
    illegal_char_pattern = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    # Write headers
    headers = ["CR", "Site", "Command", "Parameter", "Report", "TAG", "Script", "Full command log"]
    column_widths = [20, 20, 45, 25, 25, 20, 40, 70]

    for col_num, (header, width) in enumerate(zip(headers, column_widths), start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(size=12, bold=True)
        cell.fill = PatternFill(start_color="0EA1DD", end_color="0EA1DD", fill_type="solid")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

    total = len(log_data)
    for idx, (FILE_LOG, Site, type_script, line_execute, TAG_RESULT, line_full_log, att_list) in enumerate(log_data, start=2):
        log1 = illegal_char_pattern.sub('', line_full_log)
        TAG_REMARK, TAG_COLOR = CATEGORY_CHECKING(TAG_RESULT)
        line_execute = next((m.group(1) for line in att_list if (m := re.match(r"^[A-Z0-9_\-]{4,180}>(.*?)$", line))), "") if line_execute == "NULL" else line_execute

        parameter_match = re.search(r"^SET\s+.*?\s+(.*?)\s+\=", line_execute, re.IGNORECASE)
        parameter = parameter_match.group(1) if parameter_match else ""

        if not parameter:
            parameter_match = re.search(r">\s+set\s+.*?\s+(.*?)\s+", line_execute, re.IGNORECASE)
            parameter = parameter_match.group(1) if parameter_match else ""

        # Write data to cells
        ws.cell(row=idx, column=1, value=FILE_LOG).font = Font(size=9, bold=True)
        ws.cell(row=idx, column=2, value=Site).font = Font(size=9, bold=True)
        ws.cell(row=idx, column=3, value=line_execute).font = Font(size=9)
        ws.cell(row=idx, column=4, value=parameter).font = Font(size=9)
        ws.cell(row=idx, column=5, value=TAG_REMARK).font = Font(size=9)
        ws.cell(row=idx, column=6, value=log1).font = Font(size=9)
        ws.cell(row=idx, column=7, value=type_script).font = Font(size=9)
        ws.cell(row=idx, column=8, value='\n'.join(att_list)).font = Font(size=9)
        ws.cell(row=idx, column=5).fill = PatternFill(start_color=TAG_COLOR, end_color=TAG_COLOR, fill_type="solid")
        # GUI progress update
        if progress_callback is not None:
            percent = int((idx-1)/total*100)
            progress_callback(percent)

    # Create a DataFrame from the log data for pivot tables
    df = pd.DataFrame(log_data, columns=["CR", "Site", "Script", "Command", "TAG", "Full Log", "Attachments"])
    df['Report'] = df['TAG'].apply(lambda x: CATEGORY_CHECKING(x)[0])

    # Create first pivot table (by Site)
    pivot_df = pd.pivot_table(
        df,
        values='Command',
        index=['Site','Report'],
        aggfunc='count',
        fill_value=0
    )

    # Add total column
    pivot_df['Total'] = pivot_df.sum(axis=1)

    # Create pivot table sheet
    pivot_sheet = wb.create_sheet(title="Pivot Table")
    
    # Write pivot table headers
    headers = ['Site', 'Report'] + list(pivot_df.columns)
    for col_num, header in enumerate(headers, start=1):
        cell = pivot_sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(size=12, bold=True)
        cell.fill = PatternFill(start_color="0EA1DD", end_color="0EA1DD", fill_type="solid")
        pivot_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15

    # Write pivot table data
    for row_idx, ((site, report), row) in enumerate(pivot_df.iterrows(), start=2):
        pivot_sheet.cell(row=row_idx, column=1, value=site).font = Font(size=9, bold=True)
        pivot_sheet.cell(row=row_idx, column=2, value=report).font = Font(size=9, bold=True)
        for col_idx, value in enumerate(row, start=3):
            cell = pivot_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=9)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    # Add total row
    total_row = len(pivot_df) + 2
    pivot_sheet.cell(row=total_row, column=1, value="Total").font = Font(size=9, bold=True)
    for col_idx, value in enumerate(pivot_df.sum(), start=3):
        cell = pivot_sheet.cell(row=total_row, column=col_idx, value=value)
        cell.font = Font(size=9, bold=True)
        cell.fill = PatternFill(start_color="0EA1DD", end_color="0EA1DD", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Create second pivot table (Report Summary)
    report_summary = df.groupby('Report').size().reset_index(name='Count')
    total_commands = report_summary['Count'].sum()
    report_summary['Percentage'] = (report_summary['Count'] / total_commands * 100).round(1).astype(str) + '%'

    # Create Report Summary sheet
    report_sheet = wb.create_sheet(title="Pivot DF Summary")
    
    # Write headers
    headers = ['Report', 'Count', 'Percentage']
    for col_num, header in enumerate(headers, start=1):
        cell = report_sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(size=12, bold=True)
        cell.fill = PatternFill(start_color="0EA1DD", end_color="0EA1DD", fill_type="solid")
        report_sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 30

    # Write data
    for row_idx, (_, row) in enumerate(report_summary.iterrows(), start=2):
        # Write Report
        cell = report_sheet.cell(row=row_idx, column=1, value=row['Report'])
        cell.font = Font(size=9)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write Count
        cell = report_sheet.cell(row=row_idx, column=2, value=row['Count'])
        cell.font = Font(size=9)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write Percentage
        cell = report_sheet.cell(row=row_idx, column=3, value=row['Percentage'])
        cell.font = Font(size=9)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Apply color fill based on Report category
        TAG_RESULT = row['Report']
        if TAG_RESULT == "1 MOs created" or TAG_RESULT == "1 MOs set" or TAG_RESULT == "Mo deleted":
            TAG_COLOR="42FF00"
        elif re.search(r"Success to deleted", TAG_RESULT, re.IGNORECASE):
            TAG_COLOR="42FF00"    
        elif re.search(r"MOs already exists", TAG_RESULT, re.IGNORECASE):
            TAG_COLOR="FFFF00"                     
        elif re.search(r"^Total.*?MOs\s+attempted", TAG_RESULT, re.IGNORECASE):
            TAG_COLOR="42FF00"
        else:
            TAG_COLOR="FF7575"        

        # Apply color to the Report cell (first cell in the row)
        report_sheet.cell(row=row_idx, column=1).fill = PatternFill(start_color=TAG_COLOR, end_color=TAG_COLOR, fill_type="solid")

    # Add total row
    total_row = len(report_summary) + 2
    report_sheet.cell(row=total_row, column=1, value="Total").font = Font(size=9, bold=True)
    report_sheet.cell(row=total_row, column=2, value=total_commands).font = Font(size=9, bold=True)
    report_sheet.cell(row=total_row, column=3, value="100%").font = Font(size=9, bold=True)

    # Format total row
    for col in range(1, 4):
        cell = report_sheet.cell(row=total_row, column=col)
        cell.fill = PatternFill(start_color="0EA1DD", end_color="0EA1DD", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    
    
    if df_LOG_status_bf is not None and not df_LOG_status_bf.empty and df_LOG_status_af is not None and not df_LOG_status_af.empty:
        ###############################################
        # Merge the two DataFrames on 'NODENAME' and 'MO'
        cols = ['NODENAME', 'MO', 'administrativeState', 'operationalState']
        df_result = (
            pd.merge(df_LOG_status_bf[cols], df_LOG_status_af[cols], on=['NODENAME', 'MO'], suffixes=('_bf', '_af'))
            .assign(
                state_pair_bf=lambda x: x['administrativeState_bf'].astype(str) + '|' + x['operationalState_bf'].astype(str),
                state_pair_af=lambda x: x['administrativeState_af'].astype(str) + '|' + x['operationalState_af'].astype(str),
                state_match=lambda x: x['state_pair_bf'] == x['state_pair_af']
            )[
                ['NODENAME', 'MO',
                'administrativeState_bf', 'operationalState_bf',
                'administrativeState_af', 'operationalState_af',
                'state_match']
            ]
        )


        df_LOG_Alarm_af['Alarm'] = df_LOG_Alarm_af[['Severity', 'Object', 'Problem', 'Cause', 'AdditionalText']].astype(str).agg('|'.join, axis=1)
        df_LOG_Alarm_bf['Alarm'] = df_LOG_Alarm_bf[['Severity', 'Object', 'Problem', 'Cause', 'AdditionalText']].astype(str).agg('|'.join, axis=1)
        # Step 1: Add suffix
        df_LOG_Alarm_bf = df_LOG_Alarm_bf.add_suffix('_Before')
        df_LOG_Alarm_af = df_LOG_Alarm_af.add_suffix('_After')

        # Step 2: Merge
        df_compare_alarm = pd.merge(
            df_LOG_Alarm_af,
            df_LOG_Alarm_bf,
            left_on=['NODENAME_After', 'Alarm_After'],
            right_on=['NODENAME_Before', 'Alarm_Before'],
            how='left'
        )

        # Step 3: Apply Remarks
        def get_alarm_remark(row):
            try:
                if pd.notna(row['Alarm_Before']):
                    return 'Alarm Existing'
                else:
                    return 'NEW Alarm'
            except KeyError:
                return 'NO DATA BEFORE'

        df_compare_alarm['Remarks'] = df_compare_alarm.apply(get_alarm_remark, axis=1)
        df_compare_alarm = df_compare_alarm.rename(columns={
        'NODENAME_After': 'NODENAME'
        })
        # Optional: Rearrange columns for readability
        cols = [
            'NODENAME', 
            'Date_Before', 'Time_Before', 'Alarm_Before',
            'Date_After', 'Time_After', 'Alarm_After', 'Remarks'
        ]
        df_compare_alarm = df_compare_alarm[cols]




        # Save the Excel workbook
        # Write each DataFrame to a separate sheet if not empty
        def write_df_to_sheet(wb, df, sheet_name, highlight_remarks=None):
            if df is None or df.empty:
                return

            ws = wb.create_sheet(title=sheet_name)

            # Styles
            bold = Font(bold=True)
            border = Border(*(Side(style="thin") for _ in range(4)))
            fills = {
                "yellow": PatternFill("solid", fgColor="fafbbe"),
                "yellow_remark": PatternFill("solid", fgColor="fbff02"),
                "blue": PatternFill("solid", fgColor="ADD8E6"),
                "orange": PatternFill("solid", fgColor="FFD580"),
                "green": PatternFill("solid", fgColor="90EE90"),
                "red": PatternFill("solid", fgColor="FF9999"),
            }

            # Column color mapping
            color_map = {
                "Alarm": ['yellow', 'blue', 'blue', 'blue', 'orange', 'orange', 'orange', 'green'],
                "CellState": ['yellow','yellow', 'blue', 'blue', 'orange', 'orange', 'green']
            }

            remarks_idx = None
            state_match_idx = None
            for r, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                ws.append(row)

                for c, val in enumerate(row, 1):
                    cell = ws.cell(row=r, column=c)
                    if r == 1:  # Header
                        cell.font = bold
                        cell.border = border
                        color = color_map.get(sheet_name, [])[c-1] if c-1 < len(color_map.get(sheet_name, [])) else None
                        if color:
                            cell.fill = fills[color]
                        if val == "Remarks":
                            remarks_idx = c
                        if sheet_name == "CellState" and val == "state_match":
                            state_match_idx = c
                    else:
                        cell.font = Font(size=9)
                        if highlight_remarks and remarks_idx and row[remarks_idx - 1] == highlight_remarks:
                            cell.fill = fills["yellow_remark"]
                        # Style for state_match column in CellState
                        if sheet_name == "CellState" and state_match_idx and c == state_match_idx:
                            if str(val).strip().lower() == "true":
                                cell.fill = fills["green"]
                            elif str(val).strip().lower() == "false":
                                cell.fill = fills["red"]

            # Autofit columns
            for col in ws.columns:
                width = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(width + 2, 50)

        # === USAGE ===    
        write_df_to_sheet(wb, df_compare_alarm, "Alarm", highlight_remarks="NEW Alarm")
        write_df_to_sheet(wb, df_result, "CellState")
    

    wb.save(excel_filename)
     







if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = ExcelReaderApp()
    sys.exit(app.exec_())

class ExcelWriterThread(QThread):
    progress_changed = pyqtSignal(int)
    finished = pyqtSignal(str)  # Pass the output file path
    phase_changed = pyqtSignal(str)
    details_changed = pyqtSignal(str)

    def __init__(self, log_data, excel_filename, selected_file, summary_data=None, summary_filename=None, summary_mode=False):
        super().__init__()
        self.log_data = log_data
        self.excel_filename = excel_filename
        self.selected_file = selected_file
        self.summary_data = summary_data
        self.summary_filename = summary_filename
        self.summary_mode = summary_mode

    def run(self):
        try:
            print("[DEBUG] ExcelWriterThread started, setting progress to 0")
            self.progress_changed.emit(0)
            # Use the shared write_logs_to_excel function with a progress callback
            from .report_generator import write_logs_to_excel
            def debug_progress(val):
                print(f"[DEBUG] ExcelWriterThread progress: {val}")
                self.progress_changed.emit(val)
            write_logs_to_excel(self.log_data, self.excel_filename, self.selected_file, progress_callback=debug_progress)
            self.progress_changed.emit(100)
            self.finished.emit(self.excel_filename)
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            self.finished.emit("")
