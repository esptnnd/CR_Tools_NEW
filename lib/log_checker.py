# -----------------------------------------------------------------------------
# Author      : esptnnd
# Company     : Ericsson Indonesia
# Created on  : 7 May 2025
# Description : CR TOOLS by esptnnd — built for the ECT Project to help the team
#               execute faster, smoother, and with way less hassle.
#               Making life easier, one script at a time!
# -----------------------------------------------------------------------------

import os
import pandas as pd
import zipfile
import re
import time
import subprocess
from PyQt5.QtWidgets import QApplication, QProgressDialog, QMessageBox # Needed for UI elements
from PyQt5.QtCore import QTimer
import openpyxl
from openpyxl.styles import PatternFill
from .utils import debug_print


def check_logs_and_export_to_excel(parent=None, log_check_mode="Normal Log Checking"):
    import os
    import pandas as pd
    import zipfile

    download_dir = os.path.join(os.path.dirname(__file__), '..', '02_DOWNLOAD') # Adjust path to 02_DOWNLOAD
    zip_files = [fname for fname in os.listdir(download_dir) if fname.lower().endswith('.zip')]
    result_rows = []
    result_alarm_check = []  # New list for alarm data
    result_cellstatus_check = []  # New list for cell status data
    result_lte_check = []  # New list for LTE data
    result_nr_check = []  # New list for NR data
    progress = None

    # Determine logic based on log_check_mode
    if log_check_mode != "Normal Log Checking":
        is_normal_check = True
    else:
        is_normal_check = False
    ##is_normal_check = log_check_mode == "Normal Log Checking"
    ##is_mocn_check = log_check_mode == "3G_MOCN_CELL_LTE Checking"

    # Load BEFORE.xlsx if needed
    df_alarm_before = None
    if is_normal_check:
        before_path = os.path.join(os.path.dirname(__file__), '..', '00_IPDB', 'BEFORE.xlsx')
        if os.path.exists(before_path):
            try:
                if QApplication.instance() is not None:
                    progress = QProgressDialog("Loading BEFORE.xlsx...", None, 0, 1, parent)
                    progress.setWindowTitle("Loading Before Data")
                    progress.setMinimumDuration(0)
                    progress.setValue(0)
                    progress.setCancelButton(None)
                    progress.show()
                df_alarm_before = pd.read_excel(before_path, sheet_name='Alarm_Before')
                df_mobatch_status = pd.read_excel(before_path, sheet_name='Status')
                df_rnc_cell_activity = pd.read_excel(before_path, sheet_name='RNC_cell_activity')
                df_3GMOCN_cell_activity = pd.read_excel(before_path, sheet_name='3GMOCN_LTE')

                df_rnc_cell_activity = df_rnc_cell_activity.applymap(lambda x: x.strip() if isinstance(x, str) else x)                
                if progress is not None:
                    progress.setValue(1)
                    progress.close()
            except Exception as e:
                debug_print(f"Error loading BEFORE.xlsx: {e}")
                if progress is not None:
                    progress.close()
                QMessageBox.warning(parent, "Warning", f"Could not load BEFORE.xlsx: {e}")


    # Identify non-empty zip files
    loop_zip_ok_file = []
    total_zips = len(zip_files)
    if QApplication.instance() is not None:
        progress = QProgressDialog("Checking zip files...", None, 0, total_zips, parent)
        progress.setWindowTitle("Checking Zips")
        progress.setMinimumDuration(0)
        progress.setValue(0)
        progress.setCancelButton(None)
        progress.show()

    for idx, fname in enumerate(zip_files):
        zip_path = os.path.join(download_dir, fname)
        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                # Check if the zip file is empty
                if zf.namelist(): # Check if namelist is NOT empty
                    loop_zip_ok_file.append(zip_path)
                else:
                    debug_print(f"[INFO] Skipping empty zip file: {fname}")
        except zipfile.BadZipFile:
            debug_print(f"Skipping bad zip file: {fname}")
        except Exception as e:
            debug_print(f"Error processing zip file {fname} during check: {e}")

        if progress is not None:
            progress.setValue(idx + 1)
            QApplication.processEvents()

    if progress is not None:
        progress.close()

    # Now process the non-empty zip files
    if QApplication.instance() is not None:
        progress = QProgressDialog("Checking logs and exporting...", None, 0, len(loop_zip_ok_file), parent)
        progress.setWindowTitle("Checking Logs")
        progress.setMinimumDuration(0)
        progress.setValue(0)
        progress.setCancelButton(None)
        progress.show() # This might block the UI update, consider using a timer or worker for heavy tasks


    result_rows = []

    # Define check patterns for different data types
    item_check_list = [
        {
            'name': 'cellstatus',
            'start_marker': '####LOG_cellstatus',
            'end_marker': '####END_LOG_cellstatus',
            'sheet_name': 'Cell_Status',
            'result_list': []
        },
        {
            'name': 'LTE_data',
            'start_marker': '####LOG_bandwidth',
            'end_marker': '####END_LOG_bandwidth',
            'sheet_name': 'LTE_data',
            'result_list': []
        },
        {
            'name': 'NR_data',
            'start_marker': '####LOG_BAND_NR_SECTOR',
            'end_marker': '####END_LOG_BAND_NR_SECTOR',
            'sheet_name': 'NR_data',
            'result_list': []
        },
        {
            'name': 'RNC_celldata',
            'start_marker': '####LOG_CELL_3G',
            'end_marker': '####END_LOG_CELL_3G',
            'sheet_name': 'RNC_celldata',
            'result_list': []
        },
        {
            'name': 'RNC_IUBdata',
            'start_marker': '####LOG_IUB_RNC_3G',
            'end_marker': '####END_LOG_IUB_RNC_3G',
            'sheet_name': 'RNC_IUBdata',
            'result_list': []
        }        
        
    ]

    for idx, zip_path in enumerate(loop_zip_ok_file):
        fname = os.path.basename(zip_path)
        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                for member in zf.namelist():
                    # Looking for LOG/<ANY_FOLDER>/<NODENAME>.log
                    parts = member.split('/')
                    if len(parts) == 3 and parts[0] == 'LOG' and parts[2].endswith('.log'):
                        folder = parts[1]
                        nodename = parts[2][:-4]  # remove .log
                        try:
                            with zf.open(member) as f:
                                # Read lines safely, handle potential errors
                                try:
                                    # Added 'rb' mode for reading binary, and decode
                                    lines = [line.decode(errors='ignore').strip() for line in f.readlines()]
                                except Exception as read_err:
                                     debug_print(f"Error reading {member} in {fname}: {read_err}")
                                     lines = [] # Assign empty list on error

                            error1 = any('Checking ip contact...Not OK' in line for line in lines)
                            error2 = any(line.startswith('Unable to connect to ') for line in lines)
                            error3 = any('tbac control - unauthorised network element' in line for line in lines)

                            # Update: UNREMOTE if either condition is met
                            ###if any('Checking ip contact...Not OK' in line for line in lines) or any(line.startswith('Unable to connect to ') for line in lines):
                            if error1 or error2 or error3:
                                remark = 'UNREMOTE'
                            else:
                                remark = 'OK'
                            result_rows.append({
                                'FILE': fname,
                                'FOLDER': folder,
                                'NODENAME': nodename,
                                'REMARK': remark,
                                'Count': len(lines)
                            })

                            # Check for alarm logs in 99_CONCEK2 folder
                            if len(parts) == 3 and parts[0] == 'LOG' and folder == '99_Hygiene_collect' and parts[2].endswith('.log'):
                                alarm_section = False
                                # Initialize section flags for all check patterns
                                section_flags = {item['name']: False for item in item_check_list}
                                
                                for line in lines:
                                    if '####LOG_Alarm' in line:
                                        alarm_section = True
                                        continue
                                    elif '####END_LOG_Alarm' in line:
                                        alarm_section = False
                                        continue
                                    
                                    # Check for start/end markers for each pattern
                                    for item in item_check_list:
                                        if item['start_marker'] in line:
                                            section_flags[item['name']] = True
                                            continue
                                        elif item['end_marker'] in line:
                                            section_flags[item['name']] = False
                                            continue
                                    
                                    if alarm_section and line.strip() and ';' in line:
                                        try:
                                            # Split line by semicolon and strip whitespace from each part
                                            parts = [part.strip() for part in line.split(';')]
                                            if len(parts) >= 7 and parts[0] != "Date" and parts[1] != "Time":  # Skip header row and ensure we have all required fields
                                                result_alarm_check.append({
                                                    'FILE': fname,
                                                    'FOLDER': folder,
                                                    'NODENAME': nodename,
                                                    'Date': parts[0],
                                                    'Time': parts[1],
                                                    'Severity': parts[2],
                                                    'Problem': parts[4],
                                                    'Object': parts[3],
                                                    'Cause': parts[5],
                                                    'AdditionalText': parts[6]
                                                })
                                        except Exception as alarm_err:
                                            debug_print(f"Error processing alarm line in {member}: {alarm_err}")

                                    # Process data for each active section
                                    for item in item_check_list:
                                        if section_flags[item['name']] and line.strip() and ';' in line:
                                            try:
                                                # Split line by semicolon and strip whitespace from each part
                                                parts = [part.strip() for part in line.split(';')]
                                                
                                                # If this is the header row, store the column mapping
                                                if parts[0].lower() == "mo":
                                                    header_mapping = {col.lower(): idx for idx, col in enumerate(parts)}
                                                    continue
                                                
                                                # Skip empty lines
                                                if not parts[0]:
                                                    continue
                                                    
                                                # Create data entry with MO as special column
                                                data_entry = {
                                                    'FILE': fname,
                                                    'NODENAME': nodename,
                                                    'MO': parts[header_mapping.get('mo', 0)]
                                                }
                                                
                                                # Add all other columns from the header mapping
                                                for col_name, idx in header_mapping.items():
                                                    if col_name != 'mo':  # Skip MO as it's already added
                                                        data_entry[col_name] = parts[idx]
                                                
                                                item['result_list'].append(data_entry)
                                                
                                            except Exception as data_err:
                                                debug_print(f"Error processing {item['name']} line in {member}: {data_err}")

                        except Exception as open_err:
                            debug_print(f"Error opening {member} in {fname}: {open_err}")
        except zipfile.BadZipFile:
            # This should ideally not happen if we checked in the first loop,
            # but keeping for robustness.
            debug_print(f"Skipping bad zip file during processing: {fname}")
        except Exception as e:
            debug_print(f"Error processing zip file {fname}: {e}")

        if progress is not None:
            progress.setValue(idx + 1)
            QApplication.processEvents() # Allow UI updates

    if progress is not None:
        progress.close()

    # Create DataFrames
    df = pd.DataFrame(result_rows, columns=['FILE', 'FOLDER', 'NODENAME', 'REMARK', 'Count'])
    df_connection_check = df.copy()
    out_path = os.path.join(download_dir, 'MOBATCH_Check.xlsx')
    
    # Export to Excel with multiple sheets
    try:
        with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
            # Check if df is not empty before writing
            if not df.empty:
                df.to_excel(writer, sheet_name='Connection_Check', index=False)
            
            if result_alarm_check:
                df_alarm = pd.DataFrame(result_alarm_check)
                if not df_alarm.empty:
                    debug_print("SKIP EXPORT")
                    df_alarm.to_excel(writer, sheet_name='ALARM', index=False)
                
                # Compare with before data if available
                if df_alarm_before is not None and not df_alarm_before.empty:
                    try:
                        # Rename columns in df_alarm_before to add _BEFORE suffix
                        df_mobatch_status = df_mobatch_status.rename(columns={'REMARK': 'Mobatch_Before'})
                        df_mobatch_status = df_mobatch_status.drop(columns=['FILE', 'FOLDER','Count'])
                        df_alarm_before = df_alarm_before.drop(columns=['FILE', 'FOLDER'])
                        df_alarm_before = df_alarm_before.add_suffix('_BEFORE')
                        
                        # Verify required columns exist before merge
                        required_columns = ['NODENAME', 'Severity', 'Problem', 'Object']
                        if all(col in df_alarm.columns for col in required_columns) and \
                           all(col in df_alarm_before.columns for col in [f"{col}_BEFORE" for col in required_columns]):
                            
                            # Merge the dataframes on the specified keys
                            df_compare_alarm = pd.merge(
                                df_alarm,
                                df_alarm_before,
                                left_on=['NODENAME', 'Severity', 'Problem', 'Object'],
                                right_on=['NODENAME_BEFORE', 'Severity_BEFORE', 'Problem_BEFORE', 'Object_BEFORE'],
                                how='left'
                            )
                            
                            # Initialize REMARK column if it doesn't exist
                            ##if 'REMARK_COMPARE' not in df_compare_alarm.columns:
                               ##df_compare_alarm['REMARK_COMPARE'] = 'NO DATA BEFORE'
                            
                            # Check if df_mobatch_status exists and has required columns
                            if 'df_mobatch_status' in locals() and df_mobatch_status is not None and not df_mobatch_status.empty and 'NODENAME' in df_mobatch_status.columns:
                                # Merge the dataframes on df mobatch
                                df_compare_alarm = pd.merge(
                                    df_compare_alarm,
                                    df_mobatch_status,
                                    left_on=['NODENAME'],
                                    right_on=['NODENAME'],
                                    how='left'
                                )

                            # Remove specific _BEFORE columns
                            columns_to_drop = ['NODENAME_BEFORE', 'Severity_BEFORE',
                            'Object_BEFORE','Cause_BEFORE','AdditionalText_BEFORE']
                            df_compare_alarm = df_compare_alarm.drop(columns=[col for col in columns_to_drop if col in df_compare_alarm.columns])
                            
                            # Add REMARK column based on whether the row exists in before data and Status_Before
                            def get_remark(row):
                                try:
                                    if pd.notna(row['Problem_BEFORE']):
                                        return 'Alarm Existing'
                                    elif pd.isna(row['Mobatch_Before']):
                                        return 'NO DATA BEFORE'
                                    elif row['Mobatch_Before'] == 'UNREMOTE':
                                        return 'Before Site Unremote'
                                    elif row['Mobatch_Before'] == 'OK':
                                        return 'NEW Alarm'
                                    else:
                                        return 'NO DATA BEFORE'
                                except KeyError:
                                    return 'NO DATA BEFORE'

                            df_compare_alarm['REMARK_COMPARE'] = df_compare_alarm.apply(get_remark, axis=1)
                            
                            # Export the comparison data
                            df_compare_alarm.to_excel(writer, sheet_name='ALARM_COMPARE', index=False)
                            
                            try:
                                # Create yellow fill pattern
                                yellow_fill = PatternFill(start_color='FFFF00',
                                                        end_color='FFFF00',
                                                        fill_type='solid')
                                
                                # Get the worksheet and apply highlighting
                                worksheet = writer.sheets['ALARM_COMPARE']
                                
                                # Highlight NEW Alarm rows in yellow
                                for idx, row in enumerate(df_compare_alarm['REMARK_COMPARE'], start=2):
                                    if row == 'NEW Alarm':
                                        for cell in worksheet[idx]:
                                            cell.fill = yellow_fill
                                
                                debug_print("Successfully highlighted NEW Alarm rows")
                            except Exception as e:
                                debug_print(f"Warning: Could not highlight cells: {str(e)}")
                                import traceback
                                debug_print(traceback.format_exc())
                            
                            # Export the before data
                            ##df_alarm_before.to_excel(writer, sheet_name='ALARM_BEFORE', index=False)
                        else:
                            debug_print("Warning: Required columns missing for alarm comparison")
                            debug_print(f"Available columns in df_alarm: {df_alarm.columns.tolist()}")
                            debug_print(f"Available columns in df_alarm_before: {df_alarm_before.columns.tolist()}")
                    except Exception as e:
                        debug_print(f"Error during alarm comparison: {str(e)}")
                        import traceback
                        debug_print(traceback.format_exc())
            
            # Export data for each check pattern if available
            for item in item_check_list:
                if item['result_list']:
                    df_data = pd.DataFrame(item['result_list'])
                    if log_check_mode == "collect data Hygiene":
                        df_data.to_excel(writer, sheet_name=item['sheet_name'], index=False)
                    ##Cell_Status
                    if item['sheet_name'] == 'Cell_Status':
                        column_order = ['NODENAME','MO',                      
                            'administrativestate',
                            'operationalstate' ]
                        df_data = df_data.assign(**{col: pd.NA for col in column_order if col not in df_data.columns})
                        df_data = df_data[column_order]
                        df_data.to_excel(writer, sheet_name=item['sheet_name'], index=False)

                    elif (item['sheet_name'] == 'LTE_data' and 
                        'earfcndl' in df_data.columns and 'earfcnul' in df_data.columns and log_check_mode == "3G_MOCN_CELL_LTE_Checking"):

                        df_mob_check = df_connection_check.copy()
                        column_order = ['NODENAME','REMARK']                    
                        df_mob_check = df_mob_check[column_order]
                        df_mob_check = df_mob_check.rename(columns={'REMARK': 'REMARK_MOBATCH'})
                        df_3GMOCN_cell_activity = pd.merge(df_3GMOCN_cell_activity, df_mob_check, left_on=['NODENAME'], right_on=['NODENAME'], how='left')


                        df_lte_data = df_data.copy()
                        df_lte_data['MO'] = df_lte_data['MO'].str.replace('EUtranCell(FDD|TDD)=', '', case=False, regex=True)

                        column_order = ['NODENAME','MO',
                            'earfcndl',
                            'earfcnul',                        
                            'administrativestate',
                            'operationalstate' ]
                        df_lte_data = df_lte_data.assign(**{col: pd.NA for col in column_order if col not in df_lte_data.columns})
                        df_lte_data = df_lte_data[column_order]

                        df_source = df_lte_data.rename(columns={'MO': 'SOURCE_MO_CELL'})
                        df_target = df_lte_data.rename(columns={'MO': 'TARGET_MO_CELL'})

                        # Single merge operation for both source and target
                        df_merged = pd.merge(
                            pd.merge(df_3GMOCN_cell_activity, df_source, left_on=['NODENAME', 'CELLNAME SOURCE'], right_on=['NODENAME', 'SOURCE_MO_CELL'], how='left', suffixes=('', '_SOURCE')),
                            df_target,
                            left_on=['NODENAME', 'CELLNAME TARGET'],
                            right_on=['NODENAME', 'TARGET_MO_CELL'],
                            how='left',
                            suffixes=('_SOURCE', '_TARGET')
                        )

                        df_merged = df_merged.drop(columns=['SOURCE_MO_CELL', 'TARGET_MO_CELL'])

                        # Remove duplicates based on specified columns for LTE_data
                        df_merged = df_merged.drop_duplicates(subset=['NODENAME', 'CELLNAME SOURCE', 'CELLNAME TARGET'], keep='first')
                        
                        # Print information about removed duplicates
                        removed_count = len(df_3GMOCN_cell_activity) - len(df_merged)
                        if removed_count > 0:
                            debug_print(f"Removed {removed_count} duplicate entries from LTE_data")
                        
                        # Export merged data to 3G_MOCN_CELL_LTE sheet
                        df_merged.to_excel(writer, sheet_name='3G_MOCN_CELL_LTE', index=False)



                    elif item['sheet_name'] == 'RNC_celldata' and 'iublinkref' in df_data.columns and log_check_mode == "RNC_Rehoming_Checking":
                        # Remove 'UtranCell=' from MO column (case-insensitive)                    
                        df_rnc_dump = df_data.copy()
                        df_rnc_dump = df_rnc_dump[['NODENAME', 'MO']]  # Keep only NODENAME and MO columns
                        df_rnc_dump['MO'] = df_rnc_dump['MO'].str.replace('UtranCell=', '', case=False)

                        cell_status_data = pd.DataFrame(next(item['result_list'] for item in item_check_list if item['sheet_name'] == 'Cell_Status'))
                        cell_status_data['STATE'] = cell_status_data['administrativestate'].astype(str) + ' ' + cell_status_data['operationalstate'].astype(str)
                        cell_status_data['MO'] = cell_status_data['MO'].str.replace('UtranCell=', '', case=False)
                        column_order = ['NODENAME','MO','STATE']
                        cell_status_data = cell_status_data[column_order]
                        ##df_merged = pd.merge(df_merged, cell_status_data, left_on=['RNC_SOURCE', 'CELLNAME'], right_on=['SOURCE_NODE', 'IUB_SOURCE'], how='left')
                        
                        
                        # Create source and target copies of df_rnc_dump for merging ##asli nya nanti kita skip kalau udah aman
                        ##df_source = df_rnc_dump.rename(columns={'NODENAME': 'SOURCE_NODE', 'MO': 'SOURCE_MO'})
                        ##df_target = df_rnc_dump.rename(columns={'NODENAME': 'TARGET_NODE', 'MO': 'TARGET_MO'})

                        # Create source and target copies of df_rnc_dump for merging
                        df_source = cell_status_data.rename(columns={'NODENAME': 'SOURCE_NODE', 'MO': 'SOURCE_MO', 'STATE': 'SOURCE_STATE'})
                        df_target = cell_status_data.rename(columns={'NODENAME': 'TARGET_NODE', 'MO': 'TARGET_MO', 'STATE': 'TARGET_STATE'})


                        # Single merge operation for both source and target
                        df_merged = pd.merge(
                            pd.merge(df_rnc_cell_activity, df_source, left_on=['RNC_SOURCE', 'CELLNAME'], right_on=['SOURCE_NODE', 'SOURCE_MO'], how='left'),
                            df_target,
                            left_on=['RNC_TARGET', 'CELLNAME'],
                            right_on=['TARGET_NODE', 'TARGET_MO'],
                            how='left'
                        )
                        
                        # Add remarks based on merge results
                        ##df_merged['REMARK_SOURCE'] = df_merged['SOURCE_NODE'].notna().map({True: 'DEFINED', False: 'N/A'})
                        ##df_merged['REMARK_TARGET'] = df_merged['TARGET_NODE'].notna().map({True: 'DEFINED', False: 'N/A'})
                        
                        # Drop temporary columns
                        df_merged = df_merged.drop(columns=['SOURCE_NODE', 'SOURCE_MO', 'TARGET_NODE', 'TARGET_MO'])


                        # Create source and target copies of df_rnc_dump for merging 
                        ### IUBLINK CHECK
                        df_rnc_dump = df_data.copy()
                        df_rnc_dump = df_rnc_dump[['NODENAME', 'iublinkref']]    
                        df_rnc_dump['iublinkref'] = df_rnc_dump['iublinkref'].str.replace('IubLink=', '', case=False)                 
                        df_source = df_rnc_dump.rename(columns={'NODENAME': 'SOURCE_NODE', 'iublinkref': 'IUB_SOURCE'})
                        df_target = df_rnc_dump.rename(columns={'NODENAME': 'TARGET_NODE', 'iublinkref': 'IUB_TARGET'})

                        ###RNC_IUBdata
                        cell_IUB_data = pd.DataFrame(next(item['result_list'] for item in item_check_list if item['sheet_name'] == 'RNC_IUBdata'))
                        cell_IUB_data['STATE'] = cell_IUB_data['administrativestate'].astype(str) + ' ' + cell_IUB_data['operationalstate'].astype(str)
                        cell_IUB_data['MO'] = cell_IUB_data['MO'].str.replace('IubLink=', '', case=False)
                        column_order = ['NODENAME','MO','STATE']
                        cell_IUB_data = cell_IUB_data[column_order] 
                        # Create source and target copies of df_rnc_dump for merging
                        df_source = cell_IUB_data.rename(columns={'NODENAME': 'SOURCE_NODE', 'MO': 'SOURCE_MO', 'STATE': 'SOURCE_IUB_STATE'})
                        df_target = cell_IUB_data.rename(columns={'NODENAME': 'TARGET_NODE', 'MO': 'TARGET_MO', 'STATE': 'TARGET_IUB_STATE'})


                        # Single merge operation for both source and target
                        df_merged = pd.merge(
                            pd.merge(df_merged, df_source, left_on=['RNC_SOURCE', 'IUBLINK'], right_on=['SOURCE_NODE', 'SOURCE_MO'], how='left'),
                            df_target,
                            left_on=['RNC_TARGET', 'IUBLINK'],
                            right_on=['TARGET_NODE', 'TARGET_MO'],
                            how='left'
                        )
                        df_merged = df_merged.drop(columns=['SOURCE_NODE', 'SOURCE_MO', 'TARGET_NODE', 'TARGET_MO'])
                        for col in ['SOURCE_STATE', 'TARGET_STATE', 'SOURCE_IUB_STATE','TARGET_IUB_STATE']: df_merged[col] = df_merged[col].fillna("N/A") if col in df_merged.columns else df_merged.get(col)

                        # Add remarks based on merge results
                        ##df_merged['REMARK_IUB_SOURCE'] = df_merged['SOURCE_NODE'].notna().map({True: 'IUB DEFINED', False: 'N/A'})
                        ##df_merged['REMARK_IUB_TARGET'] = df_merged['TARGET_NODE'].notna().map({True: 'IUB DEFINED', False: 'N/A'})
                        # Drop temporary columns
                        ##df_merged = df_merged.drop(columns=['SOURCE_NODE', 'IUB_SOURCE', 'TARGET_NODE', 'IUB_TARGET'])

                        ####


                        ###remove duplicate
                        df_merged = df_merged.drop_duplicates(subset=['CELLNAME', 'IUBLINK', 'RNC_SOURCE', 'RNC_TARGET'], keep='first')
                                           
                                            

                        
                        # Export merged data to RNC_ACTIVITY sheet
                        df_merged.to_excel(writer, sheet_name='RNC_ACTIVITY', index=False)
                        ##df_data.to_excel(writer, sheet_name=item['sheet_name'], index=False)
               
                        

            # Adjust column widths for all sheets
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    # Find the maximum length of content in the column
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    # Set column width (max 25 characters)
                    adjusted_width = min(max_length + 2, 25)  # Add 2 for padding
                    worksheet.column_dimensions[column_letter].width = adjusted_width

    except Exception as e:
        debug_print(f"Error writing to Excel file: {str(e)}")
        import traceback
        debug_print(traceback.format_exc())
        raise

    debug_print(f"Exported check results to {out_path}")
    # Sort sheets: Connection_Check, 3G_MOCN_CELL_LTE, Cell_Status first (if they exist), then the rest
    try:
        import openpyxl
        wb = openpyxl.load_workbook(out_path)
        desired_order = ['Connection_Check', '3G_MOCN_CELL_LTE', 'Cell_Status']
        # Move desired sheets to the front in the specified order if they exist
        for sheet_name in reversed(desired_order):
            if sheet_name in wb.sheetnames:
                idx = wb.sheetnames.index(sheet_name)
                ws = wb[sheet_name]
                if sheet_name == '3G_MOCN_CELL_LTE':
                    ws.sheet_properties.tabColor = "FFFF00"  # Keep yellow tab
                wb._sheets.insert(0, wb._sheets.pop(idx))
        wb.save(out_path)
    except Exception as e:
        debug_print(f"Could not reorder sheets or set tab color: {e}")

    # Display a success message box after export
    if QApplication.instance() is not None:
        msg_box = QMessageBox(parent)
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setWindowTitle("Export Complete")
        msg_box.setText(f"Log check completed and exported to {out_path}\n DONT FORGET TO SAVE THE LOG BEFORE DOWNLOAD NEW LOGS")
        
        # If compare_before is True, open the Excel file after message box is closed

        def open_excel_file():
            try:
                # Use the appropriate command based on the operating system
                if os.name == 'nt':  # Windows
                    os.startfile(out_path)
                else:  # Linux/Mac
                    subprocess.run(['xdg-open', out_path])
            except Exception as e:
                debug_print(f"Error opening Excel file: {e}")
            
        # Connect the finished signal to open the Excel file
        msg_box.finished.connect(open_excel_file)
        
        msg_box.exec_()





        